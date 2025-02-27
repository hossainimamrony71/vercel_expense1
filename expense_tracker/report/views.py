from datetime import datetime, date, timedelta  # Import the required classes directly
from io import BytesIO

from django.shortcuts import render
from django.http import HttpResponse, HttpResponseForbidden
from django.views import View
from django.db.models import Sum
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils.decorators import method_decorator

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# Import the models you want to report on
from expense.models import MoneyAllocation, Transaction, LoanRequest, ExpenseCategory
from account.models import User

from collections import defaultdict
from datetime import datetime, timedelta, date
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from io import BytesIO

class FinanceReportView(View):
    def get(self, request):
        # Provide user types for filtering (if needed)
        user_types = [choice[0] for choice in User.CHOICES_USER_TYPE]
        return render(request, 'report.html', {'user_types': user_types})

    def post(self, request):
        # -----------------------------
        # Get form values
        # -----------------------------
        user_type = request.POST.get('user_type')  # may be empty
        model_choice = request.POST.get('model_choice')
        date_range_option = request.POST.get('date_range_option')
        custom_start_date = request.POST.get('custom_start_date')
        custom_end_date = request.POST.get('custom_end_date')

        if model_choice != 'transaction':
            return HttpResponse(
                "New report format is only available for Transactions.",
                status=400
            )

        # -----------------------------
        # Determine the date range
        # -----------------------------
        today = timezone.now().date()
        if date_range_option == 'today':
            start_date = today
            end_date = today
        elif date_range_option == 'weekly':
            days_since_saturday = (today.weekday() - 5) % 7
            start_date = today - timedelta(days=days_since_saturday)
            end_date = start_date + timedelta(days=6)
            if end_date > today:
                end_date = today
        elif date_range_option == 'monthly':
            start_date = today.replace(day=1)
            next_month = start_date.replace(day=28) + timedelta(days=4)
            computed_end_date = next_month - timedelta(days=next_month.day)
            end_date = computed_end_date if computed_end_date <= today else today
        elif date_range_option == 'yearly':
            start_date = date(today.year, 1, 1)
            computed_end_date = date(today.year, 12, 31)
            end_date = computed_end_date if computed_end_date <= today else today
        elif date_range_option == 'custom' and custom_start_date and custom_end_date:
            try:
                start_date = datetime.strptime(custom_start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(custom_end_date, '%Y-%m-%d').date()
                if end_date > today:
                    end_date = today
            except ValueError:
                return HttpResponse("Custom dates must be in YYYY-MM-DD format.", status=400)
        else:
            return HttpResponse("Invalid date range option.", status=400)

        # -----------------------------
        # Prepare the Workbook and Overall Header
        # -----------------------------
        wb = openpyxl.Workbook()
        wb.properties.creator = "Finance Report From TED & S2L"
        ws = wb.active
        ws.title = "Finance Report"

        # Overall header row (Row 1)
        ws.merge_cells('A1:E1')
        header_cell = ws['A1']
        header_cell.value = "Finance Report from TED & S2L"
        header_cell.font = Font(name="Times New Roman", size=16, bold=True)
        header_cell.alignment = Alignment(horizontal='center')

        # Filter summary (Row 2)
        ws.merge_cells('A2:E2')
        filter_text = f"Expenses | Date Range: {date_range_option.capitalize()}"
        if date_range_option == 'custom':
            filter_text += f" ({start_date} to {end_date})"
        if user_type:
            filter_text += f" | User Type: {user_type}"
        ws['A2'].value = filter_text
        ws['A2'].font = Font(italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')

        current_row = 4  # Starting row for the main report

        # -----------------------------
        # Initialize opening balance and overall category totals
        # -----------------------------
        opening_balance = self.get_opening_balance(start_date, user_type)
        overall_category_totals = {}  # For potential future use
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        summary_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # -----------------------------
        # Batch fetch transactions and money allocations for the entire date range
        # -----------------------------
        transactions_qs = Transaction.objects.filter(created_at__date__range=[start_date, end_date])
        if user_type:
            transactions_qs = transactions_qs.filter(user__user_type=user_type)
        transactions_qs = transactions_qs.order_by('created_at')
        transactions_by_date = defaultdict(list)
        for txn in transactions_qs:
            transactions_by_date[txn.created_at.date()].append(txn)

        money_alloc_qs = MoneyAllocation.objects.filter(created_at__date__range=[start_date, end_date])
        if user_type:
            money_alloc_qs = money_alloc_qs.filter(allocated_to__user_type=user_type)
        money_alloc_by_date = defaultdict(list)
        for ma in money_alloc_qs:
            money_alloc_by_date[ma.created_at.date()].append(ma)

        # -----------------------------
        # Loop through each day in the date range (never exceed today)
        # -----------------------------
        delta = end_date - start_date
        for i in range(delta.days + 1):
            current_date = start_date + timedelta(days=i)
            if current_date > today:
                break

            # Add a blank row between days (except before the first day)
            if current_row != 4:
                current_row += 1

            # Date Section Header
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            date_header_cell = ws.cell(row=current_row, column=1)
            date_header_cell.value = current_date.strftime("%A, %B %d, %Y")
            date_header_cell.font = Font(name="Garamond", size=14, bold=True)
            date_header_cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Table Headers for Transactions
            headers = ['Time', 'Particulars', 'Voucher Number', 'Method', 'Amount']
            for col_num, column_title in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = column_title
                cell.font = Font(name="Arial", bold=True)
                cell.fill = PatternFill("solid", fgColor="CCCCCC")
                cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Get transactions for the current date from the pre-fetched dictionary
            daily_transactions = transactions_by_date.get(current_date, [])
            daily_transaction_total = 0
            for txn in daily_transactions:
                col = 1
                txn_time = txn.created_at.strftime('%H:%M:%S')
                ws.cell(row=current_row, column=col, value=txn_time)
                col += 1

                particulars = ""
                if txn.category:
                    particulars = str(txn.category)
                    if txn.subcategory:
                        particulars += f" --({txn.subcategory})"
                ws.cell(row=current_row, column=col, value=particulars)
                col += 1

                ws.cell(row=current_row, column=col, value=txn.voucher)
                col += 1
                ws.cell(row=current_row, column=col, value=txn.source)
                col += 1
                amount = float(txn.ammount)
                ws.cell(row=current_row, column=col, value=amount)
                daily_transaction_total += amount

                if txn.category:
                    key = str(txn.category)
                    overall_category_totals[key] = overall_category_totals.get(key, 0) + amount
                current_row += 1

            # Subtotal Row for Transactions
            current_row += 1  # blank row for clarity
            ws.cell(row=current_row, column=1, value="Subtotal Transactions:")\
              .font = Font(name="Times New Roman", bold=True)
            ws.cell(row=current_row, column=5, value=daily_transaction_total)\
              .font = Font(name="Times New Roman", bold=True)
            current_row += 2

            # Get Money Allocations for the current date
            daily_money_allocations = money_alloc_by_date.get(current_date, [])
            daily_alloc_total = sum([ma.amount for ma in daily_money_allocations])

            # Compute Daily Totals
            total_amount = opening_balance + float(daily_alloc_total)
            closing_balance = total_amount - daily_transaction_total

            # Aggregations Section Header for Daily Financial Summary
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
            summary_header = ws.cell(row=current_row, column=1, value="Daily Financial Summary")
            summary_header.font = Font(name="Times New Roman", size=12, bold=True)
            summary_header.alignment = Alignment(horizontal='center')
            current_row += 1

            summary_data = [
                ("Opening Balance:", opening_balance),
                ("Money Allocated:", float(daily_alloc_total)),
                ("Total Amount (Opening + Allocated):", total_amount),
                ("Closing Balance:", closing_balance)
            ]
            for label, value in summary_data:
                ws.cell(row=current_row, column=1, value=label)\
                  .font = Font(name="Times New Roman", bold=True)
                ws.cell(row=current_row, column=2, value=value)\
                  .font = Font(name="Times New Roman", bold=True)
                for col in range(1, 3):
                    cell = ws.cell(row=current_row, column=col)
                    cell.fill = summary_fill
                    cell.border = thin_border
                current_row += 1

            current_row += 2  # extra spacing after the summary

            # Update opening balance for the next day
            opening_balance = closing_balance

        # -----------------------------
        # Auto–Adjust Column Widths
        # -----------------------------
        for col_cells in ws.columns:
            max_length = 0
            column_letter = None
            for cell in col_cells:
                if not isinstance(cell, MergedCell):
                    column_letter = get_column_letter(cell.column)
                    break
            if not column_letter:
                continue
            for cell in col_cells:
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value:
                        length = len(str(cell.value))
                        if length > max_length:
                            max_length = length
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width

        # -----------------------------
        # Save the workbook to a bytes buffer and return as response
        # -----------------------------
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"Transaction_Finance_Report_{start_date}_to_{end_date}.xlsx"
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    def get_opening_balance(self, date, user_type=None):
        """
        Calculate the opening balance before the given date as:
           Sum of all allocations - Sum of all transactions (up to the previous day)
        """
        allocations = MoneyAllocation.objects.filter(created_at__date__lt=date)
        transactions = Transaction.objects.filter(created_at__date__lt=date)
        if user_type:
            allocations = allocations.filter(allocated_to__user_type=user_type)
            transactions = transactions.filter(user__user_type=user_type)
        total_allocations = allocations.aggregate(total=Sum('amount'))['total'] or 0
        total_transactions = transactions.aggregate(total=Sum('ammount'))['total'] or 0
        return float(total_allocations - total_transactions)

from django.shortcuts import render
from django.db.models import Sum, Count, Case, When, DecimalField
from django.utils import timezone
from django.contrib.auth.decorators import login_required

import random  # add this at the top of your file
from django.db.models.functions import TruncMonth
from django.db.models import Sum
from dateutil.relativedelta import relativedelta
from django.db.models.functions import TruncMonth
from django.db.models import Sum
from django.utils import timezone
from django.utils.timezone import is_naive, make_aware, get_default_timezone

@login_required
def dashboard(request):
    user = request.user
    today = timezone.now().date()

    # Existing aggregations for transactions
    if user.user_type == 'admin':
        # For admin, we ignore admin’s (empty) transactions and randomly select one of TED or S2L transactions.
        chosen_type = random.choice(['ted', 's2l'])
        recent_transactions = list(
            Transaction.objects.filter(user__user_type=chosen_type).order_by('-created_at')[:10]
        )
    else:
        user_transactions = Transaction.objects.filter(user=user)
        recent_transactions = list(user_transactions.order_by('-created_at')[:10])
    
    # If needed elsewhere, you might want to compute aggregations using user_transactions.
    # For non-admin users, we can safely use their own transactions.
    # For admin, you might compute aggregations over admin-specific data or overall transactions.
    if user.user_type != 'admin':
        user_transactions = Transaction.objects.filter(user=user)
    else:
        # For admin, if you need some aggregation, you could aggregate over all transactions
        # or restrict to a specific user type. Here, we'll use all transactions as an example.
        user_transactions = Transaction.objects.all()

    aggs = user_transactions.aggregate(
        total_spending=Sum('ammount'),
        today_expense=Sum(
            Case(
                When(created_at__date=today, then='ammount'),
                default=0,
                output_field=DecimalField()
            )
        )
    )
    total_spending = aggs.get('total_spending') or 0
    today_expense = aggs.get('today_expense') or 0

    # Allocations for current user
    allocations = MoneyAllocation.objects.filter(allocated_to=user).order_by('-created_at')
    recent_allocations = list(allocations[:5])
    index_no = 2 if user.user_type == 'admin' else 1
    recent_allocated_money = allocations[:index_no].aggregate(total=Sum('amount'))['total'] or 0

    if user.user_type == 'admin':
        total_allocated = MoneyAllocation.objects.aggregate(total_allocated=Sum('amount'))['total_allocated'] or 0
        admin_allocated = MoneyAllocation.objects.filter(
            allocated_to__user_type='admin'
        ).aggregate(total_admin_allocated=Sum('amount'))['total_admin_allocated'] or 0

        total_expense = total_spending + total_allocated - admin_allocated
        total_admin_transactions = Transaction.objects.filter(user__user_type='admin').count()

        # --- New: Total Allocated to TED and S2L ---
        total_allocated_ted = MoneyAllocation.objects.filter(
            allocated_to__user_type='ted'
        ).aggregate(total_ted=Sum('amount'))['total_ted'] or 0

        total_allocated_s2l = MoneyAllocation.objects.filter(
            allocated_to__user_type='s2l'
        ).aggregate(total_s2l=Sum('amount'))['total_s2l'] or 0

        # --- New: Recent Allocations for TED and S2L ---
        ted_alloc = MoneyAllocation.objects.filter(allocated_to__user_type='ted').order_by('-created_at').first()
        s2l_alloc = MoneyAllocation.objects.filter(allocated_to__user_type='s2l').order_by('-created_at').first()
        recent_allocations_ted = int(ted_alloc.amount) if ted_alloc else 0
        recent_allocations_s2l = int(s2l_alloc.amount) if s2l_alloc else 0
    else:
        total_expense = total_spending
        total_admin_transactions = 0
        total_allocated_ted = 0
        total_allocated_s2l = 0
        recent_allocations_ted = 0
        recent_allocations_s2l = 0

    today_transactions = Transaction.objects.filter(created_at__date=today).values('user__user_type').annotate(
        count=Count('id')
    )
    counts = {item['user__user_type']: item['count'] for item in today_transactions}
    today_s2l = counts.get('s2l', 0)
    today_ted = counts.get('ted', 0)
    today_admin = counts.get('admin', 0)

    # === Donut chart data for transactions ===
    # TED transactions grouped by category
    ted_queryset = Transaction.objects.filter(user__user_type='ted').values('category__name').annotate(
        total_amount=Sum('ammount')
    ).order_by('-total_amount')
    pie_ted_labels = [item['category__name'] for item in ted_queryset]
    pie_ted_values = [float(item['total_amount'] or 0) for item in ted_queryset]

    # S2L transactions grouped by category
    s2l_queryset = Transaction.objects.filter(user__user_type='s2l').values('category__name').annotate(
        total_amount=Sum('ammount')
    ).order_by('-total_amount')[:5]
    pie_s2l_labels = [item['category__name'] for item in s2l_queryset]
    pie_s2l_values = [float(item['total_amount'] or 0) for item in s2l_queryset]
  # === Bar Chart Data for Money Allocations (Last 6 Months) ===
    # Calculate the first day of the month, six months ago (including the current month)
    six_months_ago = (today.replace(day=1) - relativedelta(months=5))
    months = []
    current_month = six_months_ago
    while current_month <= today:
        months.append(current_month)
        current_month += relativedelta(months=1)

    # Aggregate allocations for TED over the last 6 months
    ted_allocations = MoneyAllocation.objects.filter(
        allocated_to__user_type='ted',
        created_at__gte=six_months_ago
    ).annotate(month=TruncMonth('created_at')).values('month').annotate(total=Sum('amount')).order_by('month')

    # Aggregate allocations for S2L over the last 6 months
    s2l_allocations = MoneyAllocation.objects.filter(
        allocated_to__user_type='s2l',
        created_at__gte=six_months_ago
    ).annotate(month=TruncMonth('created_at')).values('month').annotate(total=Sum('amount')).order_by('month')

    def format_month(dt):
        if is_naive(dt):
            dt = make_aware(dt, get_default_timezone())
        return dt.strftime("%Y-%m")

    ted_dict = {format_month(item['month']): item['total'] for item in ted_allocations}
    s2l_dict = {format_month(item['month']): item['total'] for item in s2l_allocations}

    allocation_months = [month.strftime("%Y-%m") for month in months]
    allocation_ted = [float(ted_dict.get(month.strftime("%Y-%m"), 0)) for month in months]
    allocation_s2l = [float(s2l_dict.get(month.strftime("%Y-%m"), 0)) for month in months]

    context = {
        'balance': user.balance,
        'total_expense': total_expense,
        'today_expense': today_expense,
        'ted_admin_balance': user.admin_ted_balance,
        's2l_admin_balance': user.admin_s2l_balance,
        'recent_allocated_money': recent_allocated_money,
        'recent_transactions': recent_transactions,
        'recent_allocations': recent_allocations,
        'transactions_labels': ['S2L', 'TED'],
        'transactions_data': [today_s2l, today_ted],
        'show_row': user.user_type == 'admin',
        # Donut chart datasets:
        'pie_ted_labels': pie_ted_labels,
        'pie_ted_values': pie_ted_values,
        'pie_s2l_labels': pie_s2l_labels,
        'pie_s2l_values': pie_s2l_values,
        # --- New Context for Allocations by Type ---
        'total_allocated_ted': total_allocated_ted,
        'total_allocated_s2l': total_allocated_s2l,
        'recent_allocations_ted': recent_allocations_ted,
        'recent_allocations_s2l': recent_allocations_s2l,
        
        'allocation_months': allocation_months,
        'allocation_ted': allocation_ted,
        'allocation_s2l': allocation_s2l,
    }

    return render(request, 'admin_dashboard.html', context)



from reportlab.lib import colors
from reportlab.lib.pagesizes import A3
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from datetime import datetime
from django.http import HttpResponse
from django.db.models import Sum

import matplotlib.pyplot as plt
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib import colors
from reportlab.graphics.charts.piecharts import Pie
import math
from reportlab.lib.colors import Color

import colorsys

def generate_colors(n):
    """
    Generate a list of n eye-soothing pastel colors.
    
    Each color is generated by evenly spacing hues over [0, 1) while
    keeping a low saturation and high brightness for a soft appearance.
    """
    colors_list = []
    for i in range(n):
        # Evenly distribute hue values between 0 and 1
        hue = i / n
        
        # Lower saturation and high brightness give a pastel effect
        saturation = 0.4  # Adjust as needed for more/less vividness
        brightness = 0.95  # High brightness for a light tone
        
        # Convert HSV to RGB
        r, g, b = colorsys.hsv_to_rgb(hue, saturation, brightness)
        
        # Create a color object (assuming Color takes RGB values)
        color = Color(r, g, b)
        colors_list.append(color)
        
    return colors_list

def create_pie_chart(data, labels, width=220, height=220):
    drawing = Drawing(width, height)
    
    # Generate unique colors
    colors_list = generate_colors(len(data))
    
    # Create and configure the pie chart
    pie = Pie()
    pie.x = width / 4
    pie.y = height / 4
    pie.data = data
    pie.width = width - 40
    pie.height = height - 40
    
    total = sum(data)
    # Offset to shift pie slice labels to the left
    left_offset = 20
    for i in range(len(data)):
        pie.slices[i].fillColor = colors_list[i]
        pie.slices[i].strokeColor = colors.black
        pie.slices[i].strokeWidth = 1
        pie.slices[i].popout = 0
        
        # Calculate percentage and add it as a label
        percentage = f"{data[i] / total * 100:.1f}%"
        
        # Calculate the angle to place the label
        start_angle = sum(data[:i]) / total * 360
        end_angle = sum(data[:i+1]) / total * 360
        angle = (start_angle + end_angle) / 2
        
        # Convert angle to radians
        angle_rad = math.radians(angle)
        
        # Calculate the coordinates for the label, shifting left by left_offset
        slice_center_x = pie.x + pie.width / 2 + (pie.width / 4) * math.cos(angle_rad) - left_offset
        slice_center_y = pie.y + pie.height / 2 + (pie.height / 4) * math.sin(angle_rad)
        
        label_obj = String(slice_center_x, slice_center_y, percentage)
        label_obj.fillColor = colors.white
        label_obj.fontSize = 10
        drawing.add(label_obj)
    
    drawing.add(pie)
    
    # Add labels below the chart in a column-wise format (legend)
    # Offset to shift legend items to the left
    legend_left_offset = 40  
    label_y = height / 5 - 20  # Adjust starting vertical position
    box_width = 20
    box_height = 10
    for i, label_text in enumerate(labels):
        color = colors_list[i]
        # Calculate percentage for the label
        percentage = f"{data[i] / total * 100:.1f}%"
        label_text_with_percentage = f"{label_text} ({percentage})"
        
        # Create the colored box with adjusted x position
        color_box = Rect(10 - legend_left_offset, label_y - (i * (box_height + 5)), 
                         box_width, box_height, fillColor=color)
        drawing.add(color_box)
        
        # Create the legend text with adjusted x position
        label_string = String(40 - legend_left_offset, 
                              label_y - (i * (box_height + 5)) + box_height / 2 - 5, 
                              label_text_with_percentage)
        label_string.fillColor = colors.black
        label_string.fontSize = 10
        drawing.add(label_string)
    
    return drawing


def generate_report(request, month, year):
    current_user = request.user

    # For admin, check if a different user type was selected.
    if current_user.user_type == 'admin':
        selected_user_type = request.GET.get('report_user_type')
        if selected_user_type:
            user = User.objects.filter(user_type=selected_user_type).first()
        else:
            user = current_user
    else:
        user = current_user

    month_name = datetime(year, month, 1).strftime('%B')
    title_text = f"{user.get_user_type_display()} Monthly Report - {month_name} {year}"

    response = HttpResponse(content_type='application/pdf')
    filename = f"Financial_Report_{month_name}_{year}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Increase the bottom margin to accommodate the dynamic signature footer
    doc = SimpleDocTemplate(
        response,
        pagesize=A3,
        rightMargin=30,
        leftMargin=30,
        topMargin=30,
        bottomMargin=70
    )
    elements = []

    # Styles configuration
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name='TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        alignment=TA_CENTER,
        fontSize=18,
        textColor=colors.HexColor('#2F4F4F'),
        spaceAfter=20
    )

    section_header_style = ParagraphStyle(
        name='SectionHeader',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        alignment=TA_LEFT,
        fontSize=14,
        textColor=colors.HexColor('#2F4F4F'),
        spaceAfter=10
    )

    table_cell_style = ParagraphStyle(
        name='TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        alignment=TA_LEFT,
        fontSize=10,
        textColor=colors.black
    )

    # Table Styles for Money In and Money Out sections
    money_in_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#008000')),  # Green header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D3D3D3')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
    ])

    money_out_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF0000')),  # Red header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D3D3D3')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
    ])

    # Add report title
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 12))

    # --- Money In Section ---
    elements.append(Table([['Money In']], colWidths=[None], style=money_in_style))
    elements.append(Spacer(1, 1))

    allocations = MoneyAllocation.objects.filter(
        allocated_to=user,
        created_at__month=month,
        created_at__year=year
    )

    data_in = [['Source', 'Date', 'Amount (BDT)']]
    total_in = 0
    for alloc in allocations:
        source = alloc.source or 'N/A'
        date = alloc.created_at.strftime('%d-%b-%Y') if alloc.created_at else 'N/A'
        amount = alloc.amount or 0
        total_in += amount
        data_in.append([
            Paragraph(source, table_cell_style),
            Paragraph(date, table_cell_style),
            Paragraph(f"{amount:.2f}", table_cell_style)
        ])

    # Add total row for Money In
    data_in.append([
        Paragraph('<b>Total</b>', table_cell_style),
        Paragraph('', table_cell_style),
        Paragraph(f"<b>{total_in:.2f}</b>", table_cell_style)
    ])

    table_in = Table(data_in, colWidths=['40%', '30%', '30%'])
    table_in.setStyle(money_in_style)
    elements.append(table_in)
    elements.append(Spacer(1, 24))

    # --- Money Out Section ---
    money_out_title_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF0000')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ])

    elements.append(Table([['Money Out']], colWidths=['15%', '10%'], style=money_out_title_style))
    elements.append(Spacer(1, 1))

    transactions = Transaction.objects.filter(
        user=user,
        created_at__month=month,
        created_at__year=year
    )

    categories = transactions.values('category__name').annotate(total=Sum('ammount')).order_by('-total')
    pie_data = []
    pie_labels = []
    data_out = [['Category', 'Amount (BDT)']]
    total_out = 0

    for cat in categories:
        category_name = cat['category__name'] or 'Uncategorized'
        amount = cat['total'] or 0
        total_out += amount
        pie_data.append(float(amount))
        pie_labels.append(category_name)
        data_out.append([
            Paragraph(category_name, table_cell_style),
            Paragraph(f"{amount:.2f}", table_cell_style)
        ])

    data_out.append([
        Paragraph('<b>Total</b>', table_cell_style),
        Paragraph(f"<b>{total_out:.2f}</b>", table_cell_style)
    ])

    pie_chart = None
    if pie_data:
        pie_chart = create_pie_chart(pie_data, pie_labels)

    table_out = Table(data_out, colWidths=['60%', '40%'])
    table_out.setStyle(money_out_style)

    if pie_chart:
        combined = Table([[table_out, pie_chart]], colWidths=[doc.width * 0.6, doc.width * 0.4])
        combined.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (1, 0), (1, 0), 'CENTER'),
        ]))
        elements.append(combined)
    else:
        elements.append(table_out)

    elements.append(Spacer(1, 30))

    # --- Improved Remaining Balance Section ---
    balance = total_in - total_out
    balance_color = colors.HexColor('#2E8B57') if balance >= 0 else colors.HexColor('#B22222')

    balance_data = [
        [
            Paragraph('<b>Total Money In</b>', table_cell_style),
            Paragraph('<b>Total Money Out</b>', table_cell_style),
            Paragraph('<b>Remaining Balance</b>', table_cell_style)
        ],
        [
            Paragraph(f"{total_in:.2f}", table_cell_style),
            Paragraph(f"{total_out:.2f}", table_cell_style),
            Paragraph(f"{balance:.2f}", table_cell_style)
        ]
    ]

    # Use the same width as the left column of the money out section (60% of the document width)
    money_out_width = doc.width * 0.6
    balance_table = Table(balance_data, colWidths=[money_out_width/3.0]*3, hAlign='LEFT')
    balance_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F0F8FF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#2F4F4F')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D3D3D3')),
        ('BACKGROUND', (0, 1), (-1, 1), colors.white),
    ]))
    # Apply dynamic text color to the Remaining Balance cell (3rd column, 2nd row)
    balance_table.setStyle(TableStyle([
        ('TEXTCOLOR', (2, 1), (2, 1), balance_color)
    ]))

    elements.append(Paragraph('Remaining Balance', section_header_style))
    elements.append(balance_table)
    elements.append(Spacer(1, 30))

    # --- Signature Footer (Drawn Dynamically at the Bottom) ---
    def draw_signature_footer(canvas, doc):
        canvas.saveState()
        width, height = doc.pagesize
        footer_y = 40  # Y position from the bottom (adjust as needed)
        left_margin = doc.leftMargin
        available_width = doc.width
        column_width = available_width / 3.0

        labels = ["Account", "Checked by", "Approved by"]
        for i, label in enumerate(labels):
            x_start = left_margin + i * column_width
            x_end = x_start + column_width - 10  # Adjust gap as needed
            # Draw signature line
            canvas.setLineWidth(1)
            canvas.line(x_start, footer_y, x_end, footer_y)
            # Draw label centered below the line
            canvas.setFont("Helvetica", 10)
            canvas.drawCentredString(x_start + (column_width - 10) / 2, footer_y - 15, label)
        canvas.restoreState()

    # Build the PDF with the dynamic signature footer on every page
    doc.build(elements, onFirstPage=draw_signature_footer, onLaterPages=draw_signature_footer)
    return response



